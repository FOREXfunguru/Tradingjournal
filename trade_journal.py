import pandas as pd
import warnings
import numpy as np
import pdb
import logging
from trade import Trade
from openpyxl import load_workbook, Workbook
from pivot import PivotList
from config import CONFIG

# create logger
tj_logger = logging.getLogger(__name__)
tj_logger.setLevel(logging.INFO)

class TradeJournal(object):
    '''
    Constructor

    Class variables
    ---------------
    url: path to the .xlsx file with the trade journal
    worksheet: str, Required
               Name of the worksheet that will be used to create the object.
               i.e. trading_journal
    '''

    def __init__(self, url, worksheet):
        self.url = url
        self.worksheet = worksheet

        #read-in the 'trading_journal' worksheet from a .xlsx file into a pandas dataframe
        try:
            xls_file = pd.ExcelFile(url)
            df = xls_file.parse(worksheet, converters={'start': str, 'end': str, 'trend_i': str})
            if df.empty is True:
                raise Exception("No trades fetched for url:{0} and worksheet:{1}".format(self.url, self.worksheet))
            # replace n.a. string by NaN
            df = df.replace('n.a.', np.NaN)
            # remove trailing whitespaces from col names
            df.columns = df.columns.str.rstrip()
            self.df = df
        except FileNotFoundError:
            wb = Workbook()
            wb.create_sheet(worksheet)
            wb.save(str(self.url))

    def win_rate(self, strats):
        '''
        Calculate win rate and pips balance
        for this TradeJournal. If outcome attrb is not
        defined then it will invoke the run_trade method
        on each particular trade

        Parameters
        ----------
        strats : str
                 Comma-separated list of strategies to analyse: i.e. counter,counter_b1

        Returns
        -------
        int : number of successes
        int : number of failures
        pips : pips balance in this TradeList
        '''

        strat_l = strats.split(",")
        number_s = number_f = tot_pips = 0
        for index, row in self.df.iterrows():
            pair = row['id'].split(" ")[0]
            args = {'pair': pair}
            for c in row.keys():
                args[c] = row[c]

            t = Trade(**args)
            if t.strat not in strat_l:
                continue
            if not hasattr(t, 'outcome'):
                t.run_trade()
            if t.outcome == 'success':
                number_s += 1
            elif t.outcome == 'failure':
                number_f += 1
            tot_pips += t.pips
        tot_pips = round(tot_pips, 2)
        tot_trades = number_s+number_f
        perc_wins = round(number_s*100/tot_trades, 2)
        perc_losses = round(number_f*100/tot_trades, 2)
        print("Tot number of trades: {0}\n-------------".format(tot_trades))
        print("Win trades: {0}; Loss trades: {1}".format(number_s, number_f))
        print("% win trades: {0}; % loss trades: {1}".format(perc_wins, perc_losses))
        print("Pips balance: {0}".format(tot_pips))

        return number_s, number_f, tot_pips

    """
    def write_tradelist(self, trade_list):
        '''
        Write the TradeList to the Excel spreadsheet
        pointed by the trade_journal

        Parameters
        ----------
        tradelist : TradeList, Required

        Returns
        -------
        Nothing
        '''
        assert self.settings.has_option('trade_journal', 'worksheet_name'), \
            "'worksheet_name' needs to be defined"

        # get colnames to print in output worksheet from settings
        assert self.settings.has_option('trade_journal', 'colnames'), "'colnames' needs to be defined"
        colnames = self.settings.get('trade_journal', 'colnames').split(",")

        data = []
        for t in trade_list.tlist:
            row = []
            for a in colnames:
                if a == "session":
                    row.append(t.calc_trade_session())
                else:
                    value = None
                    try:
                        value = getattr(t, a)
                        if isinstance(value, PivotList):
                            dt_l = value.print_pivots_dates()
                            value = [d.strftime('%d/%m/%Y:%H:%M') for d in dt_l]
                    except:
                        tj_logger.warn("No value for attribute: {0}".format(a))
                        value = "n.a."
                    row.append(value)
            data.append(row)
        df = pd.DataFrame(data, columns=colnames)

        book = load_workbook(self.url)
        writer = pd.ExcelWriter(self.url, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        tj_logger.info("Creating new worksheet with trades with name: {0}".
                       format(self.settings.get('trade_journal', 'worksheet_name')))
        df.to_excel(writer, self.settings.get('trade_journal', 'worksheet_name'))
        writer.save()

    """
